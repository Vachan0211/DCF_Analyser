import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from pathlib import Path
from parser.schemas import FinancialData

DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
LIGHT_GREEN = "E2EFDA"
RED_FILL    = "FFCCCC"


def _hfont(white=True, size=11):
    return Font(name="Calibri", bold=True, size=size,
                color="FFFFFF" if white else DARK_BLUE)

def _bfont(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _cell(ws, row, col, value, fmt=None, font=None,
          fill=None, align="right", indent=0):
    c = ws.cell(row=row, column=col, value=value)
    if font:  c.font   = font
    if fill:  c.fill   = fill
    if fmt:   c.number_format = fmt
    c.alignment = Alignment(
        horizontal=align, vertical="center",
        indent=indent, wrap_text=False
    )
    c.border = _border()
    return c

def _section(ws, row, col, title, span):
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + span - 1
    )
    c = ws.cell(row=row, column=col, value=title)
    c.font      = _hfont(white=True, size=10)
    c.fill      = _fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="left",
                            vertical="center", indent=1)
    ws.row_dimensions[row].height = 17
    return c


# ── Sheet 1: DCF ─────────────────────────────────────────────
def _dcf_sheet(ws, data: FinancialData, results: dict):
    ws.title = "DCF Valuation"
    ws.sheet_view.showGridLines = False

    dcf    = results["dcf"]
    wacc_r = dcf["wacc_results"]
    projs  = dcf["projections"]

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    for col in ["D","E","F","G","H"]:
        ws.column_dimensions[col].width = 14

    # Title
    ws.merge_cells("B1:H1")
    t = ws["B1"]
    t.value = (f"DCF Valuation  —  "
               f"{data.company_info.company_name}  "
               f"({data.company_info.ticker})")
    t.font      = Font(name="Calibri", bold=True, size=15,
                       color=DARK_BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("B2:H2")
    s = ws["B2"]
    s.value = (f"Fiscal Year {data.company_info.fiscal_year}  ·  "
               f"Ended {data.company_info.fiscal_year_end_date}  ·  "
               f"$ millions")
    s.font      = Font(name="Calibri", size=10, color="808080")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    # ── WACC ──────────────────────────────────────────────────
    row = 4
    _section(ws, row, 2, "WACC Calculation", 2); row += 1

    wacc_rows = [
        ("Risk-Free Rate",         0.0411, "0.00%"),
        ("Beta",                   1.04,   "0.00"),
        ("Equity Risk Premium",    0.0446, "0.00%"),
        ("Cost of Equity (CAPM)", wacc_r["cost_of_equity"],         "0.00%"),
        ("Pre-Tax Cost of Debt",  wacc_r["pre_tax_cost_of_debt"],   "0.00%"),
        ("After-Tax Cost of Debt",wacc_r["after_tax_cost_of_debt"], "0.00%"),
        ("Weight — Equity",       wacc_r["weight_equity"],          "0.0%"),
        ("Weight — Debt",         wacc_r["weight_debt"],            "0.0%"),
        ("WACC",                  wacc_r["wacc"],                   "0.00%"),
    ]
    for ri, (label, val, fmt) in enumerate(wacc_rows):
        is_wacc = label == "WACC"
        bg = LIGHT_BLUE if is_wacc else (LIGHT_GRAY if ri%2==0 else WHITE)
        lc = _cell(ws, row, 2, label,
                   font=_bfont(bold=is_wacc),
                   fill=_fill(bg), align="left", indent=1)
        vc = _cell(ws, row, 3, val, fmt=fmt,
                   font=_bfont(bold=is_wacc), fill=_fill(bg))
        row += 1

    # ── FCF Projections ───────────────────────────────────────
    row += 1
    _section(ws, row, 2, "Free Cash Flow Projections  ($ millions)", 7)
    row += 1

    years   = [p["year"] for p in projs]
    headers = [""] + [str(y) for y in years]
    for ci, h in enumerate(headers, 2):
        c = ws.cell(row=row, column=ci, value=h if ci > 2 else "Metric")
        c.font      = _hfont(size=10)
        c.fill      = _fill(MID_BLUE)
        c.alignment = Alignment(
            horizontal="center" if ci > 2 else "left",
            vertical="center", indent=1
        )
        c.border = _border()
        ws.row_dimensions[row].height = 16
    row += 1

    fcf_rows = [
        ("Revenue",               "revenue",         '#,##0'),
        ("EBIT",                  "ebit",             '#,##0'),
        ("NOPAT",                 "nopat",            '#,##0'),
        ("D&A",                   "da",               '#,##0'),
        ("Capital Expenditures",  "capex",            '#,##0'),
        ("Change in NWC",         "change_in_nwc",    '#,##0'),
        ("Free Cash Flow",        "fcf",              '#,##0'),
        ("Discount Factor",       "discount_factor",  "0.0000"),
        ("PV of FCF",             "pv_fcf",           '#,##0'),
    ]
    for ri, (label, key, fmt) in enumerate(fcf_rows):
        highlight = key in ("fcf", "pv_fcf")
        bg = LIGHT_BLUE if highlight else (LIGHT_GRAY if ri%2==0 else WHITE)

        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = _bfont(bold=highlight)
        lc.fill      = _fill(bg)
        lc.alignment = Alignment(horizontal="left",
                                 vertical="center", indent=1)
        lc.border    = _border()

        for ci, proj in enumerate(projs, 3):
            _cell(ws, row, ci, proj[key], fmt=fmt,
                  font=_bfont(bold=highlight), fill=_fill(bg),
                  align="center")
        row += 1

    # ── Valuation Summary ─────────────────────────────────────
    row += 1
    _section(ws, row, 2, "Valuation Summary", 3); row += 1

    summary = [
        ("PV of FCFs",                dcf["sum_pv_fcf"],               "$ #,##0"),
        ("Terminal Value",            dcf["terminal_value"],            "$ #,##0"),
        ("PV of Terminal Value",      dcf["pv_terminal_value"],         "$ #,##0"),
        ("Enterprise Value",          dcf["enterprise_value"],          "$ #,##0"),
        ("Less: Net Debt",            dcf["net_debt"],                  "$ #,##0"),
        ("Equity Value",              dcf["equity_value"],              "$ #,##0"),
        ("Shares Outstanding (M)",    dcf["shares_outstanding"],        "#,##0.0"),
        ("Intrinsic Value Per Share", dcf["intrinsic_value_per_share"], "$ #,##0.00"),
    ]
    for ri, (label, val, fmt) in enumerate(summary):
        highlight = label in ("Intrinsic Value Per Share",
                              "Enterprise Value")
        bg = LIGHT_BLUE if highlight else (LIGHT_GRAY if ri%2==0 else WHITE)

        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = _bfont(bold=highlight)
        lc.fill      = _fill(bg)
        lc.alignment = Alignment(horizontal="left",
                                 vertical="center", indent=1)
        lc.border    = _border()

        vc = _cell(ws, row, 3, val, fmt=fmt,
                   font=_bfont(bold=highlight), fill=_fill(bg))
        row += 1


# ── Sheet 2: Sensitivity ──────────────────────────────────────
def _sens_sheet(ws, results: dict):
    ws.title = "Sensitivity Analysis"
    ws.sheet_view.showGridLines = False

    sens   = results["sensitivity"]
    w_list = sens["wacc_range"]
    g_list = sens["growth_range"]
    table  = sens["table"]
    base   = results["dcf"]["intrinsic_value_per_share"]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 18
    for i in range(len(w_list)):
        ws.column_dimensions[get_column_letter(i+3)].width = 13

    ws.merge_cells(
        start_row=1, start_column=2,
        end_row=1, end_column=2+len(w_list)
    )
    t = ws["B1"]
    t.value = "Sensitivity Analysis  —  Intrinsic Value per Share ($)"
    t.font  = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
    t.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(
        start_row=2, start_column=2,
        end_row=2, end_column=2+len(w_list)
    )
    s = ws["B2"]
    s.value = ("Rows = Terminal Growth Rate   ·   Columns = WACC   ·   "
               "Green = >10% above base   ·   Red = >10% below base")
    s.font  = Font(name="Calibri", size=9, color="808080", italic=True)
    ws.row_dimensions[2].height = 13

    # Header
    row = 4
    h0 = ws.cell(row=row, column=2, value="Growth \\ WACC")
    h0.font = _hfont(size=10); h0.fill = _fill(DARK_BLUE)
    h0.alignment = Alignment(horizontal="center", vertical="center")
    h0.border = _border()
    ws.row_dimensions[row].height = 18

    for ci, w in enumerate(w_list, 3):
        c = ws.cell(row=row, column=ci, value=w)
        c.font = _hfont(size=10); c.fill = _fill(DARK_BLUE)
        c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    row += 1

    for g in g_list:
        gc = ws.cell(row=row, column=2, value=g)
        gc.font = _bfont(bold=True)
        gc.fill = _fill(LIGHT_GRAY)
        gc.number_format = "0%"
        gc.alignment = Alignment(horizontal="center", vertical="center")
        gc.border = _border()
        ws.row_dimensions[row].height = 17

        for ci, w in enumerate(w_list, 3):
            row_data = table.get(round(g,2), table.get(str(round(g,2)), {}))
            val = row_data.get(round(w,2), row_data.get(str(round(w,2)), 0))
            if   val >= base * 1.10: bg = LIGHT_GREEN
            elif val <= base * 0.90: bg = RED_FILL
            else:                    bg = WHITE

            c = ws.cell(row=row, column=ci, value=val)
            c.font          = _bfont()
            c.fill          = _fill(bg)
            c.number_format = "$ #,##0.00"
            c.alignment     = Alignment(horizontal="center",
                                        vertical="center")
            c.border        = _border()
        row += 1


# ── Sheet 3: CCA ─────────────────────────────────────────────
def _cca_sheet(ws, data: FinancialData, results: dict):
    ws.title = "Comparable Companies"
    ws.sheet_view.showGridLines = False

    cca   = results["cca"]
    peers = cca["peers"]
    med   = cca["median_multiples"]
    iv    = cca["implied_value_per_share"]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 18

    ws.merge_cells("B1:G1")
    t = ws["B1"]
    t.value = "Comparable Company Analysis  (CCA)"
    t.font  = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
    t.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 24

    # Peer multiples
    row = 3
    _section(ws, row, 2, "Peer Valuation Multiples", 5); row += 1

    hdrs = ["Company", "Ticker", "EV / Revenue", "EV / EBITDA", "P / E"]
    for ci, h in enumerate(hdrs, 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _hfont(size=10); c.fill = _fill(MID_BLUE)
        c.alignment = Alignment(
            horizontal="center" if ci > 3 else "left",
            vertical="center", indent=1
        )
        c.border = _border()
        ws.row_dimensions[row].height = 16
    row += 1

    for ri, peer in enumerate(peers):
        bg = LIGHT_GRAY if ri%2==0 else WHITE
        vals = [
            (peer.get("name", peer["ticker"]), 2, "left",  "@"),
            (peer["ticker"],                   3, "center","@"),
            (round(peer.get("ev_revenue",0),1),4, "center",'0.0"x"'),
            (round(peer.get("ev_ebitda", 0),1),5, "center",'0.0"x"'),
            (round(peer.get("pe_ratio",  0),1),6, "center",'0.0"x"'),
        ]
        for val, ci, align, fmt in vals:
            c = ws.cell(row=row, column=ci, value=val)
            c.font = _bfont(); c.fill = _fill(bg)
            c.number_format = fmt
            c.alignment = Alignment(horizontal=align,
                                    vertical="center", indent=1)
            c.border = _border()
        row += 1

    # Median row
    for val, ci, align, fmt in [
        ("Median",          2, "left",  "@"),
        ("",                3, "center","@"),
        (med["ev_revenue"], 4, "center",'0.0"x"'),
        (med["ev_ebitda"],  5, "center",'0.0"x"'),
        (med["pe_ratio"],   6, "center",'0.0"x"'),
    ]:
        c = ws.cell(row=row, column=ci, value=val)
        c.font = _bfont(bold=True); c.fill = _fill(LIGHT_BLUE)
        c.number_format = fmt
        c.alignment = Alignment(horizontal=align,
                                vertical="center", indent=1)
        c.border = _border()
    row += 2

    # Implied valuation
    _section(ws, row, 2, "Implied Value Per Share  (based on peer medians)", 3)
    row += 1

    hdrs2 = ["Method", "Implied Value / Share", "vs DCF Estimate"]
    for ci, h in enumerate(hdrs2, 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _hfont(size=10); c.fill = _fill(MID_BLUE)
        c.alignment = Alignment(horizontal="center" if ci>2 else "left",
                                vertical="center", indent=1)
        c.border = _border()
        ws.row_dimensions[row].height = 16
    row += 1

    dcf_iv = results["dcf"]["intrinsic_value_per_share"]
    implied = [
        ("From EV / Revenue", iv["from_ev_revenue"]),
        ("From EV / EBITDA",  iv["from_ev_ebitda"]),
        ("From P / E",        iv["from_pe"]),
    ]
    for ri, (label, val) in enumerate(implied):
        bg = LIGHT_GRAY if ri%2==0 else WHITE
        diff = (val - dcf_iv) / dcf_iv

        ws.cell(row=row, column=2, value=label).font = _bfont()
        ws.cell(row=row, column=2).fill = _fill(bg)
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1)
        ws.cell(row=row, column=2).border = _border()

        vc = ws.cell(row=row, column=3, value=val)
        vc.font = _bfont(bold=True); vc.fill = _fill(bg)
        vc.number_format = "$ #,##0.00"
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = _border()

        dc = ws.cell(row=row, column=4, value=diff)
        dc.font = _bfont()
        dc.fill = _fill(LIGHT_GREEN if diff >= 0 else RED_FILL)
        dc.number_format = '+0.0%;-0.0%;0.0%'
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.border = _border()
        row += 1


# ── Entry point ───────────────────────────────────────────────
def generate_excel(
    data: FinancialData,
    results: dict,
    output_path: str = None
) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _dcf_sheet (wb.create_sheet("DCF Valuation"),        data, results)
    _sens_sheet(wb.create_sheet("Sensitivity Analysis"),       results)
    _cca_sheet (wb.create_sheet("Comparable Companies"), data, results)

    if output_path is None:
        output_path = f"outputs/{data.company_info.ticker}_DCF_Model.xlsx"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Excel saved to: {output_path}")
    return output_path